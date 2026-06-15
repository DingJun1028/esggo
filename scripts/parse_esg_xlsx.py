import openpyxl
import json
import os
import shutil

file_path = r"c:\Users\Administrator\OneDrive\Documents\@ESG_GO_C版_10家公司_極致擬真完整模擬填答_完全對齊A版欄位.xlsx"
temp_path = r"C:\Users\Administrator\.gemini\tmp\esggo-1\temp_copy_parse.xlsx"
output_path = r"public/c_version_mock_data.json"

if not os.path.exists(file_path):
    print("Error: Excel file does not exist at:", file_path)
    exit(1)

try:
    print("Copying Excel file to bypass lock...")
    shutil.copy2(file_path, temp_path)
    
    print("Loading workbook (data_only=True)...")
    wb = openpyxl.load_workbook(temp_path, data_only=True)
    
    # 1. Parse Sheet '01_10家公司Profile'
    companies = []
    if '01_10家公司Profile' in wb.sheetnames:
        sheet = wb['01_10家公司Profile']
        rows = list(sheet.iter_rows(values_only=True))
        headers = rows[0]
        for row in rows[1:11]: # 10 companies
            if row[0] is not None:
                company = {
                    "id": row[0],
                    "industry": row[1],
                    "name": row[2],
                    "short_name": row[3],
                    "scale": row[4],
                    "employees": row[5],
                    "revenue": row[6],
                    "locations": row[7],
                    "core_business": row[8],
                    "energy_use": row[9],
                    "electricity_kwh": row[10],
                    "water_tons": row[11],
                    "waste_tons": row[12],
                    "scope1_tco2e": row[13],
                    "scope2_tco2e": row[14]
                }
                companies.append(company)
        print(f"Parsed {len(companies)} companies profiles.")

    # 2. Parse Sheet '02_C版140題題庫'
    questions = []
    if '02_C版140題題庫' in wb.sheetnames:
        sheet = wb['02_C版140題題庫']
        rows = list(sheet.iter_rows(values_only=True))
        for row in rows[1:141]: # 140 questions
            if row[0] is not None:
                question = {
                    "id": row[0],
                    "chapter": row[1],
                    "question": row[2],
                    "input_desc": row[3],
                    "purpose": row[4],
                    "gri_ref": row[5],
                    "suggested_evidences": row[6],
                    "ai_help_desc": row[7]
                }
                questions.append(question)
        print(f"Parsed {len(questions)} C-version questions.")

    # 3. Parse Sheet '03_C版完整填答1400筆'
    answers = []
    if '03_C版完整填答1400筆' in wb.sheetnames:
        sheet = wb['03_C版完整填答1400筆']
        rows = list(sheet.iter_rows(values_only=True))
        for row in rows[1:1401]: # 1400 answers
            if row[0] is not None:
                answer = {
                    "company_id": row[0],
                    "company_type": row[1],
                    "company_name": row[2],
                    "question_id": row[3],
                    "chapter": row[4],
                    "question": row[5],
                    "content": row[6],
                    "atoms": row[7],
                    "gri_ref": row[8],
                    "suggested_evidences": row[9],
                    "ai_paragraph_direction": row[10] if len(row) > 10 else "",
                    "data_maturity": row[11] if len(row) > 11 else "",
                    "data_gap": row[12] if len(row) > 12 else ""
                }
                answers.append(answer)
        print(f"Parsed {len(answers)} filled responses (10 companies x 140 questions).")

    # Combine into a unified master database
    master_data = {
        "metadata": {
            "version": "v8.5.1-vNext",
            "title": "ESG GO C-Version 10 Companies Simulated Filled Responses Database",
            "author": "OmniAgent G4",
            "total_companies": len(companies),
            "total_questions": len(questions),
            "total_answers": len(answers)
        },
        "companies": companies,
        "questions": questions,
        "answers": answers
    }

    # Write to public/c_version_mock_data.json
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(master_data, f, ensure_ascii=False, indent=2)
    print("Successfully wrote structured JSON master database to:", output_path)

    # Clean up
    if os.path.exists(temp_path):
        os.remove(temp_path)

except Exception as e:
    print("Error parsing Excel workbook:", e)
    if os.path.exists(temp_path):
        try: os.remove(temp_path)
        except: pass
