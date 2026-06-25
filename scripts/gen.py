#!/usr/bin/env python3
import openpyxl, json, re, os, hashlib
from datetime import datetime

EXCEL = r"C:ar\www\esggo	mp_answers.xlsx"
OUTDIR = r"C:ar\www\esggoeports"
os.makedirs(OUTDIR, exist_ok=True)

wb = openpyxl.load_workbook(EXCEL, read_only=True)

# Profiles
ws = wb["01_10家公司Profile"]
profiles = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        profiles[row[0]] = dict(zip(["id","industry","name","short","scale","employees","revenue","locations","business","energy","kwh","tons"], row))

# Answers
ws2 = wb["03_C版完整填答1400筆"]
answers = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0] and row[6]:
        answers.setdefault(row[0], []).append({"qid":row[3],"ch":row[4],"ans":row[6],"gri":row[8],"atoms":row[7],"dir":row[10],"mat":row[11]})

wb.close()
print(f"Loaded {len(profiles)} profiles, {len(answers)} companies")
print(json.dumps({k: len(v) for k,v in answers.items()}))
